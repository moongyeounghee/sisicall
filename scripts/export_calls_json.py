import json
from pathlib import Path

from openpyxl import load_workbook


PROJECT = Path(__file__).resolve().parents[1]
WORKSPACE = PROJECT.parent
SOURCE = WORKSPACE / "output" / "spreadsheets" / "서울권_AI_콜_선별_개발데이터_3000건.xlsx"
OUTPUT = PROJECT / "src" / "data" / "calls3000.json"

wb = load_workbook(SOURCE, read_only=True, data_only=True)


def rows_as_dicts(sheet_name):
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = list(next(rows))
    return [dict(zip(headers, row)) for row in rows]


raw_rows = rows_as_dicts("Calls_3000")
analysis_rows = {row["call_id"]: row for row in rows_as_dicts("Route_Analysis")}

compact = []
for raw in raw_rows:
    analysis = analysis_rows[raw["call_id"]]
    compact.append(
        {
            "id": raw["call_id"],
            "requestedAt": raw["requested_at"].isoformat() if hasattr(raw["requested_at"], "isoformat") else str(raw["requested_at"]),
            "timeBucket": raw["time_bucket"],
            "originTitle": raw["pickup_name"],
            "originSub": raw["pickup_address"],
            "originDistrict": raw["pickup_district"],
            "originLat": raw["pickup_lat"],
            "originLng": raw["pickup_lng"],
            "destTitle": raw["destination_name"],
            "destSub": raw["destination_address"],
            "destRegion": raw["destination_region"],
            "destDistrict": raw["destination_district"],
            "destLat": raw["destination_lat"],
            "destLng": raw["destination_lng"],
            "pickupDistanceKm": analysis["pickup_distance_km"],
            "pickupDurationMin": analysis["pickup_duration_min"],
            "distanceKm": analysis["trip_distance_km"],
            "durationMin": analysis["trip_duration_min"],
            "homeDirectionSimilarity": analysis["home_direction_similarity"],
            "homeDistanceChangeKm": analysis["home_distance_change_km"],
            "highwayRatio": analysis["highway_ratio"],
            "destinationDemandScore": analysis["destination_demand_score"],
            "turnoverPotentialScore": analysis["turnover_potential_score"],
            "returnBurdenScore": analysis["return_burden_score"],
            "batteryUsePct": analysis["estimated_battery_use_pct"],
            "batteryAfterRidePct": analysis["battery_after_ride_pct"],
            "nearbyChargers": analysis["nearby_chargers_2km"],
            "nearbyFastChargers": analysis["nearby_fast_chargers_2km"],
            "nearestChargerDistanceKm": analysis["nearest_charger_distance_km"],
            "chargerCongestionScore": analysis["charger_congestion_score"],
            "scenarioTags": str(raw["scenario_tags"] or "").split("|"),
        }
    )

OUTPUT.write_text(json.dumps(compact, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
print(f"exported={len(compact)} path={OUTPUT} bytes={OUTPUT.stat().st_size}")
